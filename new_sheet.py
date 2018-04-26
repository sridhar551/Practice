'''from openpyxl import load_workbook
from xlutils.copy import copy as xl_copy
from xlrd import open_workbook

wb2 = load_workbook('/home/sree/Downloads/Personal/S1.xlsx')
wb = xl_copy(wb2)
wb.create_sheet('sid1')
wb.save('S1.xlsx')'''

import pandas as pd
import numpy as np
from openpyxl import load_workbook

path = r"/home/sree/Downloads/Personal/S1.xlsx"

book = load_workbook(path)
writer = pd.ExcelWriter(path, engine = 'openpyxl')
writer.book = book

x3 = np.random.randn(0,0)
df3 = pd.DataFrame(x3)

x4 = np.random.randn(0,0)
df4 = pd.DataFrame(x4)

df3.to_excel(writer, sheet_name = 'x3')
df4.to_excel(writer, sheet_name = 'x4')



writer.save()
writer.close()
