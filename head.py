'''import pandas as pd

data = pd.read_csv('/home/sree/Downloads/Personal/S3.csv', parse_dates=[0], infer_datetime_format=True)
df = pd.DataFrame(data)
df['Date_In']=pd.to_datetime(df.Date_In)
df.sort_values(by=['Date_In'])
df.to_csv('/home/sree/Downloads/Personal/S4.csv',index= False)'''

#print(df.head(10))

# - Copy and Paste Ranges using OpenPyXl library

import openpyxl

# Prepare the spreadsheets to copy from and paste too.

# File to be copied
'''wb = openpyxl.load_workbook("/home/sree/Downloads/Personal/S4.xlsx")  # Add file name
sheet = wb.get_sheet_by_name("S4")  # Add Sheet name

# File to be pasted into
template = openpyxl.load_workbook("/home/sree/Downloads/Personal/S5.xlsx")  # Add file name
temp_sheet = template.get_sheet_by_name("Sheet1")  # Add Sheet name


def copyRange(startCol, startRow, endCol=4, endRow=109, S4):
    rangeSelected = []
    # Loops through selected Rows
    for i in range(startRow, endRow + 1, 1):
        # Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol, endCol + 1, 1):
            rowSelected.append(sheet.cell(row=i, column=j).value)
        # Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)

    return rangeSelected'''
import pandas as pd

reader = pd.read_csv('/home/sree/Downloads/Personal/S4.csv')
writer = reader[reader['Date']=='2018-03-02']
writer.to_csv('/home/sree/Downloads/Personal/s5.csv', index=False)
