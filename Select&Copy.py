import openpyxl

wb = openpyxl.load_workbook("/home/sree/Downloads/Personal/Copy&Paste/data.xlsx")
sheet = wb["Sheet1"]

template = openpyxl.load_workbook("/home/sree/Downloads/Personal/Copy&Paste/data.xlsx")
temp_sheet = template["Sheet2"]

def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    for i in range(startRow, endRow+1,1):
        rowSelected = []
        for j in range(startCol, endCol+1,1):
            rowSelected.append(sheet.cell(row=i, column=j).value)
        rangeSelected.append(rowSelected)

    return rangeSelected

def pasteRange(startCol, startRow, endCol, endRow, sheetRecieving, copiedData):
    countRow = 0
    for i in range(startRow, endRow+1,1):
        countCol = 0
        for j in range(startCol, endCol+1,1):
            sheetRecieving.cell(row=i, column=j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1


#def createData():
print("Processing..........")

selectedRange = copyRange(1,425,5,530,sheet)
pastingRange = pasteRange(1,2,5,106,temp_sheet,selectedRange)
template.save("/home/sree/Downloads/Personal/Copy&Paste/data.xlsx")
print("Range copied and pasted")




